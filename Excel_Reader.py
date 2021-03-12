#! python3

#1. Read Email and copy the SmartTrack sheet to dynamic user local path                                                  ----------> 
#2. Based on the user, default select the values are pass as argv to the program/override input values  ----------> Done
#3. Take back up of BaseDataAndScoreCard.xlsx                                                                                          ----------> Done
#4. Read through the Sheet and identify the defauters                                                                               ----------> Done
#5. Scan Baseline sheet and fetch the required information for associates                                                 ----------> Done
#6. Update the Baseline sheet with Defaulter updated counter                                                                   ----------> Done
#7. Send email to Associate/escalate                                                                                                              ----------> 
#8. Remove SmartTrack sheet                                                                                                                         ----------> Done

import openpyxl
import os
import sys
import shutil
import datetime
import smtplib
from email.message import EmailMessage

####Function to Send email####
def send_email(AssociateName,AssociateEmail1,AssociateEmail2,AssociateSupervisorEmail,AssociateVSOwnerEmail,AssociateBUOwnerEmail,AssociateMonthlyctr,AssociateYTDctr):
    print('Email Sent to ' + AssociateName + AssociateEmail1 + AssociateEmail2 + AssociateSupervisorEmail + AssociateVSOwnerEmail+ AssociateBUOwnerEmail + AssociateMonthlyctr + AssociateYTDctr)


####Main Program#####
    
# Derive Working Directory
workingDir = 'C:\\trainings\\Python\\Project\\Data'

# Create Back up directory, if absent
os.chdir(workingDir)
if (os.path.isdir(workingDir+'\\backup') == False):
    os.makedirs(workingDir+'\\backup')

if (os.path.exists('SmartTrack.xlsx') == False):
    print ('SmartTrack.xlsx file is not present in working dir: ' + workingDir + '\nExiting Program..')
    sys.exit()
           
#Back up of current BaseDataAndScoreCard.xlsx in backup directory
current_ts = datetime.datetime.now()
shutil.copy(workingDir+'\\BaseDataAndScoreCard.xlsx', workingDir+'\\backup\\BaseDataAndScoreCard_'+str(current_ts.year)+str(current_ts.month)+str(current_ts.day)+'_'+str(current_ts.hour)+str(current_ts.minute)+str(current_ts.second)+'.xlsx')


workbook = openpyxl.load_workbook('SmartTrack.xlsx')
sheetname = workbook.sheetnames
sheet = workbook['Base Data']

BUNameCol = 0;
AssNameCol = 0;
VSNameCol = 0;
VSOwnerCol=0;
BUOwnerCol=0;
StatusCol=0;
LocationCol=0
WeekEndCol=0

BUName= 'Banking'
ValueStreamName= 'DFS VS ShrdSrvs'
BUOwner= 'XXX'
Location='Onsite'
VSOwner = ''

if (len(sys.argv) == 4):
    BUName = sys.argv[1]
    ValueStreamName = sys.argv[2]
    Location = sys.argv[3]
elif (len(sys.argv) == 5):
    BUName = sys.argv[1]
    ValueStreamName = sys.argv[2]
    Location = sys.argv[3]
    VSOwner = sys.argv[4]
else:
    print ('Command Line Args not available/enough. Need criteria to proceed:')
    print('Enter BU Name (Default = Banking)')
    x = input()
    if str(x) != '':
        BUName = x

    print('Enter Value Stream Name Keyword (Default = DFS VS ShrdSrvs. Enter ALL for all valuestreams in the BU)')
    x = input()
    if str(x) != '':
        ValueStreamName = x

    print('Enter Location (Default = Onsite. Enter ALL for Both Onsite & Offshore)')
    x = input()
    if str(x) != '':
        Location = x

    print('Enter VS Owner Name Keyword (Optional)')
    x = input()
    if str(x) != '':
        VSOwner = x

for i in range(1,20):
    colName = sheet.cell(row=1, column=i)
    if colName.value != None:
        if "BU" == str(colName.value.upper()):
            BUNameCol = i
        if "ASSOCIATE NAME" == str(colName.value.upper()):
            AssNameCol = i
        if ("VALUE"in str(colName.value.upper())) and ("STREAM"in str(colName.value.upper())):
            VSNameCol = i
        if (("VALUE STREAM"in str(colName.value.upper())) or ("VS"in str(colName.value.upper())) ) and ("OWNER"in str(colName.value.upper())):
            VSOwnerCol = i
        if ("BU"in str(colName.value.upper()) ) and ("OWNER"in str(colName.value.upper())):
            BUOwnerCol = i
        if "STATUS" == str(colName.value.upper()):
            StatusCol = i
        if "LOCATION" == str(colName.value.upper()):
            LocationCol = i            
        if ("WEEK"in str(colName.value.upper())) or ("ENDING"in str(colName.value.upper())):
            WeekEndCol = i
            
index = 0
AssociateEmail1=''
AssociateEmail2=''
AssociateMonthlyctr=0
AssociateYTDctr=0
AssociateSupervisorEmail = ''
AssociateVSOwnerEmail = ''
AssociateBUOwnerEmail = ''
DayofMonth=0

tempWB = openpyxl.Workbook()
sh2 = tempWB["Sheet"]

baseWorkbook = openpyxl.load_workbook('BaseDataAndScoreCard.xlsx')
baseWorksheetNames = baseWorkbook.sheetnames
baseWorksheet = baseWorkbook['Sheet']

# Identified List of Defaulters, and load in temp workbook
print('Defaulters List'.rjust(50) +'\n' +'***************'.rjust(50)+'\n'+ 'Srl_Num|AssociateName|ValueStream|Location|Weekending|AssociateEmail|SupervisorEmail|VS Owner Email|BU Owner Email|MnthlyDefCnt|YTDDefCnt')
for row in sheet.iter_rows():
    if (row[BUNameCol-1].value != BUName):
        continue
    if ((ValueStreamName.upper() != 'ALL') and (ValueStreamName not in row[VSNameCol-1].value)):
        continue    
    if ((Location.upper() != 'ALL') and (Location not in row[LocationCol-1].value)):
        continue
    if ((VSOwner != '') and (VSOwner.upper() not in row[VSOwnerCol-1].value.upper())):
        continue          

    # Defaulter matches the input criteria
    index=index+1

    # Fetch the name of the Defaulter to compare against BaseDataseheet
    defrowname=row[AssNameCol-1].value.upper().replace(',','').split()
    defrowname.sort()

    # Default Values before looping through BaseDataSheet
    AssociateEmail1='Not Available'
    AssociateEmail2='Not Available'
    AssociateMonthlyctr='Not Available'
    AssociateYTDctr='Not Available'
    AssociateSupervisorEmail = 'Not Available'
    AssociateVSOwnerEmail = 'Not Available'
    AssociateBUOwnerEmail = 'Not Available'

    for baseRow in baseWorksheet.iter_rows():
        baserowname=baseRow[3].value.upper().replace(',','').split()
        baserowname.sort()
        if (defrowname == baserowname):
            AssociateEmail1 =str(baseRow[4].value)
            AssociateEmail2 =str(baseRow[5].value)
            AssociateMonthlyctr = str(baseRow[9].value)
            AssociateYTDctr = str(baseRow[10].value)
            AssociateSupervisorEmail = str(baseRow[6].value)
            AssociateVSOwnerEmail = str(baseRow[7].value)
            AssociateBUOwnerEmail = str(baseRow[8].value)

            #Update the Monthly and YTD Counter for Associate
            DayofMonth = int(row[WeekEndCol-1].value.day)
            if DayofMonth <= 7:
                AssociateMonthlyctr = '1'
            else:
                AssociateMonthlyctr = str(int(AssociateMonthlyctr)+1)
            AssociateYTDctr = str(int(AssociateYTDctr)+1)
            baseRow[9].value = AssociateMonthlyctr
            baseRow[10].value = AssociateYTDctr

            #Send Email to Defaulter
            send_email(str(baseRow[3].value), AssociateEmail1, AssociateEmail2, AssociateSupervisorEmail, AssociateVSOwnerEmail, AssociateBUOwnerEmail, AssociateMonthlyctr, AssociateYTDctr)
            

    print(str(index)+'|'+str(row[AssNameCol-1].value)+'|'+str(row[VSNameCol-1].value)+'|'+str(row[LocationCol-1].value)+'|'+str(row[WeekEndCol-1].value).strip(' 00:00:00')+'|'+AssociateEmail1+'|'+AssociateSupervisorEmail+'|'+AssociateVSOwnerEmail+'|'+AssociateBUOwnerEmail+'|'+AssociateMonthlyctr+'|'+AssociateYTDctr)
    
    sh2.append((cell.value for cell in row))


tempWB.save("Selected_Defauters_List_current_execution.xlsx")
baseWorkbook.save("BaseDataAndScoreCard.xlsx")

#Moving the SmartTrack.xlsx file to backup directory ------------------- Commented out for timebeing -----------
#shutil.move(workingDir+'\\SmartTrack.xlsx', workingDir+'\\backup\\SmartTrack_'+str(current_ts.year)+str(current_ts.month)+str(current_ts.day)+'_'+str(current_ts.hour)+str(current_ts.minute)+str(current_ts.second)+'.xlsx')
print('End of Program')

